#!/usr/bin/env python3
"""
LV Portal Validation Formatter — Streamlit Web UI Version
=========================================================
Converts LV Portal validation exports (LLDP Mismatches, Interface Down, Optics, FEC)
into a clean multi-tab Excel report with:
  • L&R labels + patch panel lookup from cutsheet(s)
  • History/recurring flags (when previous report provided)
  • Separate tabs for T0-T1 links and Compute (T0,Host) links
  • Summary dashboard + Ghost host detection
  • Color-coded tabs and professional borders/grouping

Run:
    pip install streamlit openpyxl
    streamlit run lv_portal_formatter_streamlit.py

Same core functions and output format as the original Tkinter/console version.
"""

import sys
import os
import re
import copy
import json
import time
import tempfile
import shutil
import io
import contextlib
from pathlib import Path
from datetime import datetime
from collections import Counter

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours (same as original) ────────────────────────────────────────────────
WHITE   = "FFFFFF"
YELLOW  = "FFFF00"
LOG_BG  = "EADCF8"
HDR_BG  = "1F4E79"
HDR_FG  = "FFFFFF"
SRC_BG  = "FCE4D6"
D1_BG   = "FFF2CC"
D2_BG   = "E2F0D9"
DEST_BG = "D9EAF7"
Z_BG    = "DDEBF7"
ACT_BG  = "FFC7CE"
EXP_BG  = "C6EFCE"
LR_BG   = "FFFFFF"
LR_LOG  = "FFFFFF"
PP_BG   = "FCE4D6"
PD_BG   = "FFF2CC"
TAB_MISS = "FF0000"
TAB_DOWN = "FFA500"
TAB_OPT  = "9933FF"
TAB_FEC  = "0070C0"

def fill(h):  return PatternFill("solid", fgColor=h)
def font(color="000000", bold=False, sz=9, italic=False):
    return Font(bold=bold, italic=italic, color=color, name="Arial", size=sz)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")
def vcenter(): return Alignment(horizontal="left", vertical="center")

# ── Config (kept for compatibility, but unused in Streamlit) ─────────────────
CONFIG_FILE = os.path.expanduser("~/.lv_portal_config.json")

def load_config():
    try:
        with open(CONFIG_FILE) as f: return json.load(f)
    except: return {}

def save_config(cfg):
    try:
        with open(CONFIG_FILE, 'w') as f: json.dump(cfg, f, indent=2)
    except: pass

# ── Cutsheet loaders (UNCHANGED) ─────────────────────────────────────────────
def build_lookup(paths):
    """Build T1 reverse lookup from one or more cutsheets."""
    if isinstance(paths, str): paths = [paths]
    t0     = {}
    t1     = {}
    t1_rev = {}
    t0_to_pp = {}
    for path in paths:
        wb = load_workbook(path, read_only=True)
        sheet = next((wb[n] for n in wb.sheetnames if 'installation' in n.lower()), wb[wb.sheetnames[0]])
        count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9: continue
            lbl    = str(row[0]  or '').strip()
            dev_a  = str(row[1]  or '').strip()
            rack_a = str(row[2]  or '').strip()
            src    = str(row[3]  or '').strip()
            dmarc1 = str(row[4]  or '').strip()
            dmarc2 = str(row[5]  or '').strip()
            dest   = str(row[6]  or '').strip()
            dev_b  = str(row[7]  or '').strip()
            rack_b = str(row[8]  or '').strip()
            t1_lbl = str(row[10] or '').strip() if len(row) > 10 else ''

            if dev_a and lbl and re.match(r'\d+[LR]$', lbl):
                parts = dev_a.split()
                if len(parts) == 2:
                    k = (parts[0], parts[1])
                    t0[k] = lbl; t1[k] = t1_lbl
                    t0_to_pp[k] = {
                        'source_port': src, 'dmarc1': dmarc1, 'dmarc2': dmarc2,
                        'dest_port': dest, 'rack_b': rack_b, 't1_lbl_pp': t1_lbl,
                    }

            if dev_b and ' ' in dev_b:
                parts = dev_b.split()
                if len(parts) == 2:
                    t1_rev[(parts[0], parts[1])] = {
                        't0_lbl': lbl, 'rack_a': rack_a, 'source_port': src,
                        'dmarc1': dmarc1, 'dmarc2': dmarc2, 'dest_port': dest,
                        'rack_b': rack_b, 't1_lbl': t1_lbl,
                    }
                    count += 1
        wb.close()
        print(f"  Loaded: {os.path.basename(path)} ({count} T1 entries)")
    return t0, t1, t1_rev, t0_to_pp

def is_compute(name):
    return 'compute' in str(name or '').lower()

def cs_lookup(compute_lookup, host, port):
    if not compute_lookup: return {}
    cs = compute_lookup.get((host, port), {})
    if cs: return cs
    import re as _csl
    m = _csl.match(r'(slot\d+/port\d+-)(\d+)', port)
    if m:
        lane = int(m.group(2))
        partner = lane - 1 if lane % 2 == 0 else lane + 1
        cs = compute_lookup.get((host, f"{m.group(1)}{partner}"), {})
        if cs: return cs
        for p2 in ['1','2','3','4']:
            if p2 not in (str(lane), str(partner)):
                cs = compute_lookup.get((host, f"{m.group(1)}{p2}"), {})
                if cs: return cs
    fb = compute_lookup.get('_host_fallback', {})
    return fb.get(host, {})

def compute_port_group(compute_host, compute_port):
    import re as _cpg
    m = _cpg.match(r'(slot\d+/port\d+-)(\d+)', str(compute_port or ''))
    if not m: return (compute_host, compute_port)
    lane = int(m.group(2))
    pair_base = lane if lane % 2 == 1 else lane - 1
    return (compute_host, f"{m.group(1)}{pair_base}")

def parse_rack(rack_str):
    s = str(rack_str or '').strip()
    if ':' in s:
        parts = s.split(':')
        return f"Rack {parts[0]}", f"U{parts[1]}"
    return s, ''

def get_t0_labels(host, iface, t0, t1):
    key = (host, iface)
    t0_lbl = t0.get(key, '')
    t1_lbl = t1.get(key, '')
    is_phys = bool(t0_lbl)
    if not is_phys:
        m = re.match(r'(swp\d+)s(\d+)', iface)
        if m:
            base, lane = m.group(1), int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                pk = (host, f"{base}s{partner}")
                t0_lbl = t0.get(pk, '')
                t1_lbl = t1.get(pk, '')
    return t0_lbl, t1_lbl, is_phys

def get_mismatch_info(act_host, act_iface, t1_rev):
    mi = t1_rev.get((act_host, act_iface), {})
    if not mi:
        m = re.match(r'(swp\d+)s(\d+)', act_iface)
        if m:
            base, lane = m.group(1), int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                mi = t1_rev.get((act_host, f"{base}s{partner}"), {})
    return mi

# ── Border helpers (UNCHANGED) ───────────────────────────────────────────────
def draw_pair_borders(ws, lr_col=2, iface_col=1):
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")
    def group_key(row):
        lr  = ws.cell(row, lr_col).value or ''
        iface = str(ws.cell(row, iface_col).value or '')
        m = re.match(r'(swp\d+)s\d+', iface)
        base = m.group(1) if m else iface
        return (lr, base)
    dr = 2
    while dr <= ws.max_row:
        key = group_key(dr)
        grp_end = dr
        while grp_end + 1 <= ws.max_row and group_key(grp_end + 1) == key and key[0]:
            grp_end += 1
        for rr in range(dr, grp_end + 1):
            is_top = (rr == dr); is_bot = (rr == grp_end)
            for cc in range(1, ws.max_column + 1):
                ws.cell(rr, cc).border = Border(
                    top    = thick if is_top else thin,
                    bottom = thick if is_bot else Side(style=None),
                    left   = thick if cc == 1 else thin,
                    right  = thick if cc == ws.max_column else thin,
                )
        dr = grp_end + 1

def draw_compute_borders(ws, host_col, port_col):
    thin  = Side(style="thin",   color="AAAAAA")
    thick = Side(style="medium", color="555555")
    def key(row):
        h = str(ws.cell(row, host_col).value or '')
        p = str(ws.cell(row, port_col).value or '')
        return compute_port_group(h, p)
    dr = 2
    while dr <= ws.max_row:
        k = key(dr)
        grp_end = dr
        while grp_end + 1 <= ws.max_row and key(grp_end + 1) == k and k[1]:
            grp_end += 1
        for rr in range(dr, grp_end + 1):
            is_top = (rr == dr); is_bot = (rr == grp_end)
            for cc in range(1, ws.max_column + 1):
                ws.cell(rr, cc).border = Border(
                    top    = thick if is_top else thin,
                    bottom = thick if is_bot else Side(style=None),
                    left   = thick if cc == 1 else thin,
                    right  = thick if cc == ws.max_column else thin,
                )
        dr = grp_end + 1

def draw_row_borders(ws):
    thin   = Side(style="thin",   color="AAAAAA")
    medium = Side(style="medium", color="555555")
    for rr in range(2, ws.max_row + 1):
        for cc in range(1, ws.max_column + 1):
            ws.cell(rr, cc).border = Border(
                top    = medium if rr == 2 else thin,
                bottom = medium if rr == ws.max_row else thin,
                left   = medium if cc == 1 else thin,
                right  = medium if cc == ws.max_column else thin,
            )

# ── Header / cell helpers (UNCHANGED) ────────────────────────────────────────
def write_header_row(ws, headers, widths):
    for col, ((label, bg), w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, col)
        c.value = label; c.fill = fill(bg)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        c.alignment = center()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

def write_data_cell(ws, row, col, value, bg="FFFFFF", bold=False, sz=9, align=None):
    c = ws.cell(row, col)
    c.value = value
    c.fill  = fill(bg)
    c.font  = Font(bold=bold, color="000000", name="Arial", size=sz)
    c.alignment = align or center()

# ── Core processing (UNCHANGED except minor print tweaks) ────────────────────
def process_lldp(ws_src, t0, t1, t1_rev, t0_to_pp=None):
    t0_to_pp = t0_to_pp or {}
    raw = []
    for row in range(2, ws_src.max_row + 1):
        dev_a      = str(ws_src.cell(row, 1).value or '').strip()
        dev_b_port = str(ws_src.cell(row, 2).value or '').strip()
        dev_b_rack = str(ws_src.cell(row, 3).value or '').strip()
        exp_dev_b  = str(ws_src.cell(row, 4).value or '').strip()
        dev_a_rack = str(ws_src.cell(row, 5).value or '').strip()
        exp_rack_b = str(ws_src.cell(row, 6).value or '').strip()
        dev_b_name = str(ws_src.cell(row, 7).value or '').strip()
        dev_a_port = str(ws_src.cell(row, 8).value or '').strip()
        status     = str(ws_src.cell(row, 9).value or '').strip()
        exp_port_b = str(ws_src.cell(row, 10).value or '').strip()

        if is_compute(dev_a) or is_compute(dev_b_name) or is_compute(exp_dev_b):
            continue
        if not dev_a or not dev_a_port:
            continue

        rack, elev = parse_rack(dev_a_rack)
        exp_rack, exp_elev = parse_rack(exp_rack_b)
        act_rack, act_elev = parse_rack(dev_b_rack)

        t0_lbl, t1_lbl, is_phys = get_t0_labels(dev_a, dev_a_port, t0, t1)

        pp_own = {'source_port':'','dmarc1':'','dmarc2':'','dest_port':'','rack_b':'','t1_lbl':''}
        if is_phys and exp_dev_b and exp_port_b:
            mi_exp = t1_rev.get((exp_dev_b, exp_port_b), {})
            if not mi_exp:
                m = re.match(r'(swp\d+)s(\d+)', exp_port_b)
                if m:
                    base, lane = m.group(1), int(m.group(2))
                    partner = {0:1,1:0,2:3,3:2}.get(lane)
                    if partner is not None:
                        mi_exp = t1_rev.get((exp_dev_b, f"{base}s{partner}"), {})
            pp_own.update(mi_exp)

        mi = {}
        is_down = (status == 'INTERFACE_DOWN')
        if not is_down and is_phys and dev_b_name and dev_b_name != 'Unknown' and dev_b_port and dev_b_port != 'Unknown':
            mi = get_mismatch_info(dev_b_name, dev_b_port, t1_rev)

        raw.append({
            'host':       dev_a, 'iface': dev_a_port, 'rack': rack, 'elev': elev,
            't0_lbl': t0_lbl, 't1_lbl': t1_lbl, 'is_phys': is_phys,
            'row_type': 'downlink' if is_down else 'mismatch', 'status': status,
            'exp_host': exp_dev_b, 'exp_port': exp_port_b, 'exp_rack': exp_rack, 'exp_elev': exp_elev,
            'act_host': dev_b_name, 'act_port': dev_b_port, 'act_rack': act_rack, 'act_elev': act_elev,
            'source_port': pp_own.get('source_port', ''), 'dmarc1': pp_own.get('dmarc1', ''),
            'dmarc2': pp_own.get('dmarc2', ''), 'dest_port': pp_own.get('dest_port', ''),
            'rack_b': pp_own.get('rack_b', ''), 't1_lbl_pp': pp_own.get('t1_lbl', ''),
            'mi': mi,
        })

    # Second pass for logical rows
    pp_by_partner = {}
    mi_by_partner = {}
    for rd in raw:
        if not rd['is_phys']: continue
        m = re.match(r'(swp\d+)s(\d+)', rd['iface'])
        if m:
            lane = int(m.group(2))
            partner = {0:1,1:0,2:3,3:2}.get(lane)
            if partner is not None:
                key = (rd['host'], f"{m.group(1)}s{partner}")
                if rd['source_port']:
                    pp_by_partner[key] = {k: rd[k] for k in ['source_port','dmarc1','dmarc2','dest_port','rack_b','t1_lbl_pp']}
                if rd['mi']:
                    mi_by_partner[key] = rd['mi']

    for rd in raw:
        key = (rd['host'], rd['iface'])
        if not rd['is_phys'] and key in pp_by_partner:
            rd.update(pp_by_partner[key])
        elif not rd['is_phys'] and not rd['source_port']:
            mi_iface = re.match(r'(swp\d+)s(\d+)', rd['iface'])
            if mi_iface:
                base_i, lane_i = mi_iface.group(1), int(mi_iface.group(2))
                partner_i = {0:1,1:0,2:3,3:2}.get(lane_i)
                if partner_i is not None:
                    ck = (rd['host'], f"{base_i}s{partner_i}")
                    ct_pp = t0_to_pp.get(ck, {})
                    if ct_pp:
                        rd.update(ct_pp)
            if not rd['source_port'] and rd['exp_host'] and rd['exp_port']:
                em = re.match(r'(swp\d+)s(\d+)', rd['exp_port'])
                if em:
                    elane = int(em.group(2))
                    epartner = {0:1,1:0,2:3,3:2}.get(elane)
                    if epartner is not None:
                        pk = (rd['exp_host'], f"{em.group(1)}s{epartner}")
                        mi_exp = t1_rev.get(pk, {})
                        if mi_exp:
                            rd['source_port'] = mi_exp.get('source_port','')
                            rd['dmarc1']      = mi_exp.get('dmarc1','')
                            rd['dmarc2']      = mi_exp.get('dmarc2','')
                            rd['dest_port']   = mi_exp.get('dest_port','')
                            rd['rack_b']      = mi_exp.get('rack_b','')
                            rd['t1_lbl_pp']   = mi_exp.get('t1_lbl','')
        if not rd['is_phys'] and key in mi_by_partner:
            rd['mi'] = mi_by_partner[key]

    miss_rows = [r for r in raw if r['row_type'] == 'mismatch']
    down_rows = [r for r in raw if r['row_type'] == 'downlink']
    return miss_rows, down_rows

def get_prev_issues_lv(report_path):
    try:
        wb = load_workbook(report_path, read_only=True)
    except Exception as e:
        print(f"  Warning: could not load previous report: {e}")
        return set(), set(), set(), {}
    ws = next((wb[n] for n in wb.sheetnames if 'lldp' in n.lower() or 'mismatch' in n.lower()), None)
    prev_miss = set(); prev_down = set(); prev_rack_map = {}
    if ws:
        hc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Name'), None)
        pc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Port'), None)
        sc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='LLDP Status'), None)
        rc = next((c for c in range(1,ws.max_column+1) if str(ws.cell(1,c).value or '').strip()=='Device A Rack'), None)
        if hc and pc and sc:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[hc-1] or '').strip()
                i = str(row[pc-1] or '').strip()
                st = str(row[sc-1] or '').strip()
                rack_raw = str(row[rc-1] or '').strip() if rc else ''
                rack, _ = parse_rack(rack_raw)
                if not h or not i: continue
                if is_compute(h): continue
                if st == 'INTERFACE_DOWN': prev_down.add((h,i))
                else:                      prev_miss.add((h,i))
                prev_rack_map[(h,i)] = rack
    ws_opt = next((wb[n] for n in wb.sheetnames if 'optic' in n.lower()), None)
    prev_opt = set()
    if ws_opt:
        dn = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Device Name'), None)
        dp = next((c for c in range(1,ws_opt.max_column+1) if str(ws_opt.cell(1,c).value or '').strip()=='Device Port'), None)
        if dn and dp:
            for row in ws_opt.iter_rows(min_row=2, values_only=True):
                if not row: continue
                h = str(row[dn-1] or '').strip(); i = str(row[dp-1] or '').strip()
                if h and i and not is_compute(h): prev_opt.add((h,i))
    wb.close()
    print(f"  Previous: {len(prev_miss)} mismatches, {len(prev_down)} downlinks, {len(prev_opt)} optics")
    return prev_miss, prev_down, prev_opt, prev_rack_map

def get_history_flag(host, iface, current_type, prev_miss, prev_down, prev_opt, rack=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    rack_num = (rack.replace("Rack ","").split()[0] if rack else None)
    def _in(s):
        if not s or not isinstance(s, (set, frozenset)): return False
        try:
            return (rack_num, iface) in s or (host, iface) in s
        except TypeError:
            return False
    if current_type == 'mismatch':
        if _in(prev_miss): return "🔁 Recurring mismatch",  "FF6B6B"
        if _in(prev_down): return "⬆️ Was downlink",        "FFB347"
    elif current_type == 'downlink':
        if _in(prev_down): return "🔁 Recurring downlink",  "FF6B6B"
        if _in(prev_opt):  return "⚡ Was optic error",      "D35400"
        if _in(prev_miss): return "⬇️ Was mismatch",        "FFB347"
    elif current_type == 'optic':
        if _in(prev_opt):  return "🔁 Recurring optic",     "FF6B6B"
        if _in(prev_down): return "⬆️ Was downlink",        "FFB347"
        if _in(prev_miss): return "⬇️ Was mismatch",        "FFB347"
    return "", ""

# ── GPU / Compute cutsheet loader (UNCHANGED) ────────────────────────────────
def build_compute_lookup(paths):
    import re as _fbr
    if isinstance(paths, str): paths = [paths]
    lookup = {}
    for path in paths:
        print(f"    Opening {os.path.basename(path)}...")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        count = 0
        hdr = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for c_idx, val in enumerate(row, start=1):
                hdr[str(val or '').strip()] = c_idx
        print(f"    Headers found: {list(hdr.keys())[:8]}...")
        if 'PortA' in hdr and 'PortB' in hdr:
            ci = {k: hdr.get(k, 0) - 1 for k in
                  ['DeviceA','PortA','DeviceA Physical Port','RackA','OHR','FDF',
                   'T0','DeviceB','PortB','RackB','DeviceB Physical Port']}
            def _gv(row_vals, key):
                idx = ci.get(key, -1)
                if idx < 0 or idx >= len(row_vals): return ''
                return str(row_vals[idx] or '').strip()
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                if not row_vals: continue
                comp_host = _gv(row_vals, 'DeviceA')
                comp_port = _gv(row_vals, 'PortA')
                t0_host   = _gv(row_vals, 'DeviceB')
                t0_iface  = _gv(row_vals, 'PortB')
                if not comp_host or not t0_host or not t0_iface: continue
                entry = {
                    'lr':        _gv(row_vals, 'DeviceB Physical Port'),
                    'nic_pos':   _gv(row_vals, 'DeviceA Physical Port'),
                    'comp_host': comp_host, 'comp_port': comp_port,
                    'comp_rack': _gv(row_vals, 'RackA'),
                    'ohr':       _gv(row_vals, 'OHR'),
                    'fdf':       _gv(row_vals, 'FDF'),
                    't0_pp':     _gv(row_vals, 'T0'),
                    't0_rack':   _gv(row_vals, 'RackB'),
                    't0_host':   t0_host, 't0_iface': t0_iface,
                }
                lookup[(t0_host, t0_iface)] = entry
                lookup[(comp_host, comp_port)] = entry
                count += 1
        else:
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                if not row_vals or len(row_vals) < 36: continue
                t0_col     = str(row_vals[3]  or '').strip()
                comp_dev_a = str(row_vals[26] or '').strip()
                rack_a     = str(row_vals[27] or '').strip()
                ohr        = str(row_vals[28] or '').strip()
                fdf        = str(row_vals[29] or '').strip()
                t0_pp      = str(row_vals[30] or '').strip()
                rack_b     = str(row_vals[32] or '').strip()
                nic_pos    = str(row_vals[34] or '').strip()
                lr_label   = str(row_vals[35] or '').strip()
                if not t0_col or ' ' not in t0_col: continue
                t0_host, t0_iface = t0_col.split()[0], t0_col.split()[1]
                comp_host = comp_dev_a.split()[0] if ' ' in comp_dev_a else comp_dev_a
                comp_port = comp_dev_a.split()[-1] if ' ' in comp_dev_a else ''
                entry = {
                    'lr': lr_label, 'nic_pos': nic_pos,
                    'comp_host': comp_host, 'comp_port': comp_port,
                    'comp_rack': rack_a, 'ohr': ohr, 'fdf': fdf,
                    't0_pp': t0_pp, 't0_rack': rack_b,
                    't0_host': t0_host, 't0_iface': t0_iface,
                }
                lookup[(t0_host, t0_iface)] = entry
                if comp_host and comp_port:
                    lookup[(comp_host, comp_port)] = entry
                count += 1
        wb.close()
        print(f"  Loaded GPU cutsheet: {os.path.basename(path)} ({count} entries)")
    host_fallback = {}
    for (h, p), v in list(lookup.items()):
        if not isinstance(p, str): continue
        rack = v.get('comp_rack', '')
        if is_compute(h) and _fbr.match(r'Rack \d+ U\d+', rack):
            host_fallback[h] = v
    lookup['_host_fallback'] = host_fallback
    return lookup

# ── Sheet builders (UNCHANGED) ───────────────────────────────────────────────
def build_mispatches_sheet(wb_out, rows, prev_miss=None, prev_down=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set()
    if not rows: return
    ws = wb_out.create_sheet("Mispatches")
    ws.sheet_properties.tabColor = TAB_MISS
    headers = [
        ("Interface",        HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Elevation", HDR_BG),
        ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("DMARC2", "375623"), ("Destination_port", "17375E"),
        ("Z Interface", "17375E"), ("Z L&R", "17375E"), ("Z Rack", "17375E"), ("Z Elevation", "17375E"),
        ("Act. Interface", "9C0006"), ("Act. Rack", "9C0006"), ("Act. Elevation", "9C0006"),
        ("Exp. Interface", "375623"), ("Exp. Rack", "375623"), ("Exp. Elevation", "375623"),
        ("History", "595959"),
    ]
    widths = [12,6,10,6, 30,28,28,30, 12,6,10,6, 12,10,6, 12,10,6, 22]
    write_header_row(ws, headers, widths)
    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']
        bg = "FFFFFF"
        lr_bg = LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(rd['host'], rd['iface'], 'mismatch', prev_miss, prev_down, set())
        hist_bg = hist_col if hist_flag else bg
        vals = [
            rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
            rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
            rd['exp_port'], rd['t1_lbl'], rd['exp_rack'], rd['exp_elev'],
            rd['act_port'], rd['act_rack'], rd['act_elev'],
            rd['exp_port'], rd['exp_rack'], rd['exp_elev'],
            hist_flag,
        ]
        all_bgs = [bg,lr_bg,bg,bg, bg,bg,bg,bg, bg,lr_bg,bg,bg,
                   ACT_BG,ACT_BG,ACT_BG, EXP_BG,EXP_BG,EXP_BG, hist_bg]
        for col, (val, cell_bg) in enumerate(zip(vals, all_bgs), start=1):
            c = ws.cell(r_idx, col)
            c.value = val; c.fill = fill(cell_bg)
            fg = WHITE if (col == len(vals) and hist_flag) else "000000"
            c.font = Font(bold=(col==2 or (col==len(vals) and hist_flag)), color=fg, name="Arial", size=9)
            c.alignment = center()
    draw_pair_borders(ws, lr_col=2)

def build_downlinks_sheet(wb_out, rows, prev_miss=None, prev_down=None, prev_opt=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    if not rows: return
    ws = wb_out.create_sheet("Downlinks")
    ws.sheet_properties.tabColor = TAB_DOWN
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Elevation", HDR_BG),
        ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("DMARC2", "375623"), ("Destination_port", "17375E"),
        ("Z Interface", "17375E"), ("Z L&R", "17375E"), ("Z Rack", "17375E"), ("Z Elevation", "17375E"),
        ("History", "595959"),
    ]
    widths = [12,6,10,6, 30,28,28,30, 12,6,10,6, 22]
    write_header_row(ws, headers, widths)
    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']
        bg = "FFFFFF"; lr_bg = LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(rd['host'], rd['iface'], 'downlink', prev_miss, prev_down, prev_opt)
        hist_bg = hist_col if hist_flag else bg
        vals = [
            rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
            rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
            rd['exp_port'], rd['t1_lbl'], rd['exp_rack'], rd['exp_elev'],
            hist_flag,
        ]
        all_bgs = [bg, lr_bg, bg, bg, bg, bg, bg, bg, bg, lr_bg, bg, bg, hist_bg]
        for col, (val, cell_bg) in enumerate(zip(vals, all_bgs), start=1):
            c = ws.cell(r_idx, col)
            c.value = val; c.fill = fill(cell_bg)
            c.font = Font(bold=(col==2), color="000000", name="Arial", size=9)
            c.alignment = center()
    draw_pair_borders(ws, lr_col=2)

def parse_pp_matrix(pp_str):
    if not pp_str or pp_str.strip().upper() == 'N/A': return {}
    parts = [p.strip() for p in pp_str.split('•')]
    if len(parts) < 6: return {}
    t1_host_port = parts[6].strip() if len(parts) > 6 else ''
    t1_host = ''; t1_iface = ''
    if ' ' in t1_host_port:
        t1_host  = t1_host_port.split()[0]
        t1_iface = t1_host_port.split()[-1]
    return {
        'rack_a': parts[1] if len(parts) > 1 else '',
        'source_port': parts[2] if len(parts) > 2 else '',
        'dmarc1': parts[3] if len(parts) > 3 else '',
        'dmarc2': parts[4] if len(parts) > 4 else '',
        'dest_port': parts[5] if len(parts) > 5 else '',
        't1_host_port': t1_host_port, 't1_host': t1_host, 't1_iface': t1_iface,
        'rack_b': parts[7] if len(parts) > 7 else '',
    }

def build_optics_sheet(wb_out, ws_src, t0, t1, t1_rev, downlink_set, t0_to_pp=None, prev_miss=None, prev_down=None, prev_opt=None):
    t0_to_pp = t0_to_pp or {}
    prev_miss = prev_miss or set(); prev_down = prev_down or set(); prev_opt = prev_opt or set()
    if not ws_src: return
    ncols = ws_src.max_column
    col_port = col_rx = col_dev = col_tx = col_pp = None
    for c in range(1, ncols+1):
        hv = str(ws_src.cell(1,c).value or '').strip()
        if hv == 'Device Port':         col_port = c
        elif hv == 'Rx Power':          col_rx   = c
        elif hv == 'Device Name':       col_dev  = c
        elif hv == 'Tx Power':          col_tx   = c
        elif hv == 'Patch Panel Matrix':col_pp   = c
    if not col_port or not col_dev:
        print(f"  Optics: could not find required columns"); return

    opt_rows = []
    for row in range(2, ws_src.max_row + 1):
        port     = str(ws_src.cell(row, col_port).value or '').strip()
        rx_power = str(ws_src.cell(row, col_rx).value   or '').strip() if col_rx else ''
        dev_name = str(ws_src.cell(row, col_dev).value  or '').strip()
        tx_power = str(ws_src.cell(row, col_tx).value   or '').strip() if col_tx else ''
        pp_str   = str(ws_src.cell(row, col_pp).value   or '').strip() if col_pp else ''
        if not re.match(r'swp', port): continue
        if not port or not dev_name: continue
        if is_compute(dev_name) or is_compute(pp_str): continue
        if pp_str.upper() == 'N/A' or not pp_str: continue
        pp = parse_pp_matrix(pp_str)
        if not pp.get('source_port'): continue
        t0_lbl, _, is_p = get_t0_labels(dev_name, port, t0, t1)
        is_dl = (dev_name, port) in downlink_set
        t1_lbl_z = ''
        t1_h = pp.get('t1_host',''); t1_i = pp.get('t1_iface','')
        if t1_h and t1_i:
            t1_lbl_z = t1_rev.get((t1_h, t1_i), {}).get('t1_lbl', '')
            if not t1_lbl_z:
                m2 = re.match(r'(swp\d+)s(\d+)', t1_i)
                if m2:
                    partner2 = {0:1,1:0,2:3,3:2}.get(int(m2.group(2)))
                    if partner2 is not None:
                        t1_lbl_z = t1_rev.get((t1_h, f"{m2.group(1)}s{partner2}"), {}).get('t1_lbl', '')
        opt_rows.append({
            'host': dev_name, 'iface': port, 't0_lbl': t0_lbl, 'is_phys': is_p,
            'rx_power': rx_power, 'tx_power': tx_power,
            'rack_a': pp.get('rack_a',''), 'source_port': pp.get('source_port',''),
            'dmarc1': pp.get('dmarc1',''), 'dmarc2': pp.get('dmarc2',''),
            'dest_port': pp.get('dest_port',''), 't1_iface': t1_i, 't1_lbl_z': t1_lbl_z,
            'rack_b': pp.get('rack_b',''), 'is_dl': is_dl,
        })

    pp_base = {}
    for rd in opt_rows:
        if rd['source_port']:
            m = re.match(r'(swp\d+)s\d+', rd['iface'])
            if m: pp_base[(rd['host'], m.group(1))] = {
                k: rd[k] for k in ['rack_a','source_port','dmarc1','dmarc2','dest_port','t1_iface','t1_lbl_z','rack_b']}
    for rd in opt_rows:
        if not rd['source_port']:
            m = re.match(r'(swp\d+)s\d+', rd['iface'])
            if m:
                k = (rd['host'], m.group(1))
                if k in pp_base: rd.update(pp_base[k])

    if not opt_rows: return
    ws = wb_out.create_sheet("Optics")
    ws.sheet_properties.tabColor = TAB_OPT
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Rx Power", "7030A0"),
        ("Source_port", "C0504D"), ("DMARC1", "7F6000"), ("DMARC2", "375623"), ("Destination_port", "17375E"),
        ("Z Interface", "17375E"), ("Z L&R", "17375E"), ("Z Rack", "17375E"),
        ("DL Flag", "595959"), ("History", "595959"),
    ]
    widths = [12,6,14, 30, 30,28,28,30, 12,6,14, 24,22]
    write_header_row(ws, headers, widths)
    for out_row, rd in enumerate(opt_rows, start=2):
        ws.row_dimensions[out_row].height = 15
        is_dl = rd['is_dl']
        row_bg = "C8C8C8" if is_dl else "FFFFFF"
        dl_flag = "⬇️ Also Downlink — skip" if is_dl else ''
        txt_fg = "888888" if is_dl else "000000"
        vals = [rd['iface'], rd['t0_lbl'], rd['rack_a'], rd['rx_power'],
                rd['source_port'], rd['dmarc1'], rd['dmarc2'], rd['dest_port'],
                rd['t1_iface'], rd['t1_lbl_z'], rd['rack_b'], dl_flag, '']
        bgs = [row_bg,"FFFFFF",row_bg, row_bg, row_bg,row_bg,row_bg,row_bg, row_bg,"FFFFFF",row_bg,
               "C8C8C8" if is_dl else row_bg, row_bg]
        for col, (val, bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(out_row, col)
            c.value = val; c.fill = fill(bg)
            c.font = Font(color=txt_fg, name="Arial", size=9)
            c.alignment = center()
    draw_pair_borders(ws, lr_col=2)

def build_fec_sheet(wb_out, ws_src, t0, t1, downlink_set):
    if not ws_src: return
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = TAB_FEC
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("Rack", HDR_BG), ("Elevation", HDR_BG),
        ("Pre-FEC BER", "7030A0"), ("Lock Status", "7030A0"), ("Z Interface", "17375E"),
        ("DL Flag", "595959"), ("History", "595959"),
    ]
    widths = [12,6,10,6, 20,40, 12, 24,22]
    write_header_row(ws, headers, widths)
    out_row = 2
    for row in range(2, ws_src.max_row + 1):
        port       = str(ws_src.cell(row, 1).value or '').strip()
        ber        = str(ws_src.cell(row, 2).value or '').strip()
        remote_dev = str(ws_src.cell(row, 3).value or '').strip()
        remote_if  = str(ws_src.cell(row, 4).value or '').strip()
        dev_rack   = str(ws_src.cell(row, 5).value or '').strip()
        dev_name   = str(ws_src.cell(row, 6).value or '').strip()
        lock_status= str(ws_src.cell(row, 7).value or '').strip()
        if not port or not dev_name: continue
        if is_compute(dev_name) or is_compute(remote_dev): continue
        t0_lbl, _, is_p = get_t0_labels(dev_name, port, t0, t1)
        rack, elev = parse_rack(dev_rack)
        is_dl = (dev_name, port) in downlink_set
        row_bg = "C8C8C8" if is_dl else "FFFFFF"
        ws.row_dimensions[out_row].height = 15
        dl_flag = "⬇️ Also Downlink — skip" if is_dl else ''
        vals = [port, t0_lbl, rack, elev, ber, lock_status, remote_if, dl_flag, '']
        bgs  = [row_bg,"FFFFFF",row_bg,row_bg, row_bg,row_bg, row_bg,
                "C8C8C8" if is_dl else row_bg, row_bg]
        for col, (val, bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(out_row, col)
            c.value = val; c.fill = fill(bg)
            c.font = Font(color="888888" if is_dl else "000000", name="Arial", size=9)
            c.alignment = center()
        out_row += 1

# ── Ghost / Compute processing (UNCHANGED) ───────────────────────────────────
GHOST_THRESHOLD = 16

def get_ghost_hosts(ws_src):
    from collections import Counter
    counts = Counter()
    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    exp_b_col = hdr.get('Expected Device B Name', 4)
    for row in range(2, ws_src.max_row + 1):
        exp_b = str(ws_src.cell(row, exp_b_col).value or '').strip()
        cur_b_col = hdr.get('Current Device B Name', exp_b_col - 3)
        cur_b = str(ws_src.cell(row, cur_b_col).value or '').strip()
        if is_compute(exp_b) and cur_b.lower() != 'missing':
            counts[exp_b] += 1
    ghosts = {host for host, cnt in counts.items() if cnt >= GHOST_THRESHOLD}
    if ghosts:
        print(f"  Ghost hosts detected ({GHOST_THRESHOLD}+ errors): {ghosts}")
    return ghosts

def _parse_location(loc_str):
    import re as _re
    m = _re.match(r'[^:]+:(\d+):(\d+)', str(loc_str or '').strip())
    if m: return f"Rack {m.group(1)}", f"U{m.group(2)}"
    rack_m = _re.search(r'rackNumber: *(\d+)', str(loc_str or ''))
    elev_m = _re.search(r'rackElevation: *(\d+)', str(loc_str or ''))
    rack = f"Rack {rack_m.group(1)}" if rack_m else ''
    elev = f"U{elev_m.group(1)}"     if elev_m else ''
    return rack, elev

def process_interface_down(ws_src, compute_lookup=None):
    import re
    compute_lookup = compute_lookup or {}
    if not ws_src: return [], [], set()
    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    c_remote_dev  = hdr.get('Remote Device Name', 1)
    c_remote_port = hdr.get('Remote Device Port', 2)
    c_src_dev     = hdr.get('Source Device Name', 3)
    c_src_loc     = hdr.get('Source Device Location', 4)
    c_src_port    = hdr.get('Source Device Port', 5)
    from collections import Counter
    counts = Counter()
    for row in range(2, ws_src.max_row + 1):
        remote = str(ws_src.cell(row, c_remote_dev).value or '').strip()
        if is_compute(remote): counts[remote] += 1
    ghost_hosts = {h for h,c in counts.items() if c >= GHOST_THRESHOLD}
    if ghost_hosts:
        print(f"  Ghost hosts detected ({GHOST_THRESHOLD}+ errors): {ghost_hosts}")
    real_rows = []; ghost_rows = []
    for row in range(2, ws_src.max_row + 1):
        exp_b    = str(ws_src.cell(row, c_remote_dev).value  or '').strip()
        comp_port= str(ws_src.cell(row, c_remote_port).value or '').strip()
        dev_a    = str(ws_src.cell(row, c_src_dev).value     or '').strip()
        loc_a    = str(ws_src.cell(row, c_src_loc).value     or '').strip()
        t0_iface = str(ws_src.cell(row, c_src_port).value    or '').strip()
        if not exp_b or not dev_a: continue
        if not is_compute(exp_b): continue
        t0_rack, t0_elev = _parse_location(loc_a)
        t0_lbl, _, is_phys = get_t0_labels(dev_a, t0_iface, {}, {})
        cs = cs_lookup(compute_lookup, dev_a, t0_iface)
        if not cs.get('ohr') and comp_port:
            cs_rev = cs_lookup(compute_lookup, exp_b, comp_port)
            if cs_rev.get('ohr'):
                cs = {**cs_rev, **{k:v for k,v in cs.items() if v}}
        if not cs.get('comp_rack'):
            fb = compute_lookup.get('_host_fallback', {})
            cs_fb = fb.get(exp_b, {})
            if cs_fb.get('comp_rack'): cs = {**cs, 'comp_rack': cs_fb['comp_rack']}
        if cs.get('t0_rack'):
            t0_rack = cs['t0_rack']
        import re as _rlre2
        _rls2 = _rlre2.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rls2: t0_rack, t0_elev = _rls2.group(1), _rls2.group(2)
        comp_rack_full = cs.get('comp_rack', '')
        import re as _cr2re
        _crm = _cr2re.match(r'Rack (\d+) U(\d+)', comp_rack_full)
        comp_rack = f"Rack {_crm.group(1)}" if _crm else comp_rack_full
        comp_elev = f"U{_crm.group(2)}"      if _crm else ''
        if cs.get('lr'): t0_lbl = cs['lr']
        import re as _dre2
        _dm2 = _dre2.match(r'slot\d+/port\d+-(\d+)', comp_port)
        comp_is_phys = (int(_dm2.group(1)) % 2 == 1) if _dm2 else True
        rd = {
            'host': dev_a, 'iface': t0_iface, 't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            'exp_host': exp_b, 'comp_port': comp_port,
            'comp_rack': comp_rack, 'comp_elev': comp_elev,
            'cur_b': 'interface down',
            'comp_is_phys': comp_is_phys,
            'nic_pos':  cs.get('nic_pos', ''), 'ohr': cs.get('ohr', ''),
            'fdf': cs.get('fdf', ''), 't0_pp': cs.get('t0_pp', ''),
        }
        if exp_b in ghost_hosts:
            ghost_rows.append(rd)
        else:
            real_rows.append(rd)
    return real_rows, ghost_rows, ghost_hosts

def process_compute_lldp(ws_src, compute_lookup=None):
    import re
    compute_lookup = compute_lookup or {}
    ghost_hosts = get_ghost_hosts(ws_src)
    real_rows = []; ghost_rows = []
    hdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    new_fmt = 'Device A Name' in hdr
    if new_fmt:
        col_dev_a   = hdr.get('Device A Name', 1)
        col_loc_a   = hdr.get('Device A Location', 2)
        col_port_a  = hdr.get('Device A Port', 3)
        col_cur_b   = hdr.get('Current Device B Name', 4)
        col_exp_b   = hdr.get('Expected Device B Name', 7)
        col_exp_loc = hdr.get('Expected B Location', 8)
        col_exp_port= hdr.get('Expected Device B Port', 9)
        col_pp      = hdr.get('Patch Panel Matrix', 10)
        col_err     = hdr.get('Error Message', 11)
    else:
        col_dev_a=1; col_loc_a=3; col_port_a=2; col_cur_b=8
        col_exp_b=4; col_exp_loc=6; col_exp_port=5; col_pp=None; col_err=10
    for row in range(2, ws_src.max_row + 1):
        dev_a   = str(ws_src.cell(row, col_dev_a).value or '').strip()
        loc_a   = str(ws_src.cell(row, col_loc_a).value or '').strip()
        port_a  = str(ws_src.cell(row, col_port_a).value or '').strip()
        cur_b   = str(ws_src.cell(row, col_cur_b).value or '').strip()
        exp_b   = str(ws_src.cell(row, col_exp_b).value or '').strip()
        exp_loc = str(ws_src.cell(row, col_exp_loc).value or '').strip()
        exp_port= str(ws_src.cell(row, col_exp_port).value or '').strip()
        err_msg = str(ws_src.cell(row, col_err).value or '').strip()
        if not dev_a: continue
        if not (is_compute(exp_b) or is_compute(cur_b)): continue
        if cur_b.lower() == 'missing': continue
        t0_iface = port_a if re.match(r'swp', port_a) else (
            re.search(r'(swp\d+s\d+)', err_msg).group(1)
            if re.search(r'(swp\d+s\d+)', err_msg) else 'Unknown')
        t0_rack, t0_elev = _parse_location(loc_a)
        comp_port = exp_port
        comp_rack, comp_elev = _parse_location(exp_loc)
        t0_lbl, _, is_phys = get_t0_labels(dev_a, t0_iface, {}, {})
        cs = cs_lookup(compute_lookup, dev_a, t0_iface)
        if not cs.get('ohr') and comp_port:
            cs_rev = cs_lookup(compute_lookup, exp_b, comp_port)
            if cs_rev.get('ohr'):
                cs = {**cs_rev, **{k:v for k,v in cs.items() if v}}
        if not cs.get('comp_rack'):
            fb = compute_lookup.get('_host_fallback', {})
            cs_fb = fb.get(exp_b, {})
            if cs_fb.get('comp_rack'): cs = {**cs, 'comp_rack': cs_fb['comp_rack']}
        if cs:
            t0_lbl    = cs.get('lr',      t0_lbl)
            is_phys   = bool(compute_lookup.get((dev_a, t0_iface)))
            t0_rack   = cs.get('t0_rack', t0_rack)
            comp_rack = cs.get('comp_rack', comp_rack)
            comp_port = cs.get('comp_port', comp_port)
        import re as _rlre
        _rls = _rlre.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rls: t0_rack, t0_elev = _rls.group(1), _rls.group(2)
        _rlc = _rlre.match(r'(Rack \d+) (U\d+)', comp_rack)
        if _rlc: comp_rack, comp_elev = _rlc.group(1), _rlc.group(2)
        import re as _dre
        _dm = _dre.match(r'slot\d+/port\d+-(\d+)', comp_port)
        comp_is_phys = (int(_dm.group(1)) % 2 == 1) if _dm else True
        rd = {
            'host': dev_a, 'iface': t0_iface, 't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            'exp_host': exp_b, 'comp_port': comp_port,
            'comp_rack': comp_rack, 'comp_elev': comp_elev,
            'cur_b': cur_b,
            'comp_is_phys': comp_is_phys,
            'nic_pos':  cs.get('nic_pos', ''), 'ohr': cs.get('ohr', ''),
            'fdf': cs.get('fdf', ''), 't0_pp': cs.get('t0_pp', ''),
        }
        if exp_b in ghost_hosts:
            ghost_rows.append(rd)
        else:
            real_rows.append(rd)
    return real_rows, ghost_rows, ghost_hosts

def build_compute_sheet(wb_out, rows, tab_name="Downlinks", tab_colour="70AD47",
                        prev_miss=None, prev_down=None):
    prev_miss = prev_miss or set(); prev_down = prev_down or set()
    if not rows: return
    ws = wb_out.create_sheet(tab_name)
    ws.sheet_properties.tabColor = tab_colour
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("T0 Rack", HDR_BG), ("Elevation", HDR_BG),
        ("OHR", "7F6000"), ("FDF", "375623"), ("T0 PP", "17375E"),
        ("Compute Host", "375623"), ("Compute Port", "375623"), ("NIC Position", "833C00"),
        ("Compute Rack", "375623"), ("Compute U", "375623"),
        ("Current Device", "9C0006"), ("History", "595959"),
    ]
    widths = [12, 6, 12, 6, 28, 28, 30, 40, 16, 14, 12, 8, 40, 22]
    write_header_row(ws, headers, widths)
    rows = sorted(rows, key=lambda r: (
        r.get('exp_host',''),
        compute_port_group(r.get('exp_host',''), r.get('comp_port','')),
        r.get('comp_port','')
    ))
    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']; bg="FFFFFF"; lr_bg=LR_BG if p else LR_LOG
        hist_flag, hist_col = get_history_flag(
            rd['host'], rd['iface'], 'mismatch', prev_miss, prev_down, set())
        hist_bg = hist_col if hist_flag else bg
        comp_p = rd.get('comp_is_phys', True)
        comp_port_bg = "E2F0D9" if comp_p else "D5F5E3"
        vals = [rd['iface'], rd['t0_lbl'], rd['rack'], rd['elev'],
                rd.get('ohr',''), rd.get('fdf',''), rd.get('t0_pp',''),
                rd['exp_host'], rd['comp_port'], rd.get('nic_pos',''), rd['comp_rack'],
                rd.get('comp_elev',''), rd['cur_b'], hist_flag]
        all_bgs = [bg,lr_bg,bg,bg, "FFF2CC","E2F0D9","D9EAF7",
                   "E2F0D9",comp_port_bg,"FDDCB5","E2F0D9","E2F0D9",
                   ACT_BG, hist_bg]
        for col,(val,cell_bg) in enumerate(zip(vals,all_bgs),start=1):
            c=ws.cell(r_idx,col); c.value=val; c.fill=fill(cell_bg)
            fg=WHITE if (col==len(vals) and hist_flag) else "000000"
            c.font=Font(bold=(col==2 or (col==len(vals) and hist_flag)),color=fg,name="Arial",size=9)
            c.alignment=center()
    draw_compute_borders(ws, host_col=8, port_col=9)

def build_compute_optics_sheet(wb_out, ws_src, ghost_hosts, t0, t1, compute_lookup=None):
    compute_lookup = compute_lookup or {}
    if not ws_src: return
    rows = []
    import re
    _ohdr = {str(ws_src.cell(1,c).value or '').strip(): c for c in range(1, ws_src.max_column+1)}
    _col_remote_dev  = _ohdr.get('Remote Device Name', 1)
    _col_remote_port = _ohdr.get('Remote Device Port', 2)
    _col_src_dev     = _ohdr.get('Source Device Name', 3)
    _col_src_loc     = _ohdr.get('Source Device Location', _ohdr.get('Source Location', 5))
    _col_src_port    = _ohdr.get('Source Device Port', _ohdr.get('Source Port', 4))
    _col_rx          = _ohdr.get('Rx Power', 6)
    for row in range(2, ws_src.max_row + 1):
        remote_dev  = str(ws_src.cell(row, _col_remote_dev).value  or '').strip()
        remote_port = str(ws_src.cell(row, _col_remote_port).value or '').strip()
        src_dev     = str(ws_src.cell(row, _col_src_dev).value     or '').strip()
        src_port    = str(ws_src.cell(row, _col_src_port).value    or '').strip()
        src_loc     = str(ws_src.cell(row, _col_src_loc).value     or '').strip()
        rx_power    = str(ws_src.cell(row, _col_rx).value          or '').strip()
        if not src_dev or not src_port: continue
        if not is_compute(remote_dev): continue
        if remote_dev in ghost_hosts: continue
        t0_rack, t0_elev = _parse_location(src_loc)
        t0_lbl, _, is_phys = get_t0_labels(src_dev, src_port, t0, t1)
        import re as _opr
        comp_cs = cs_lookup(compute_lookup, remote_dev, remote_port)
        comp_rack_full = comp_cs.get('comp_rack', '')
        _cr = _opr.match(r'Rack (\d+) U(\d+)', comp_rack_full)
        comp_rack_num = f"Rack {_cr.group(1)}" if _cr else ''
        comp_u        = f"U{_cr.group(2)}"      if _cr else ''
        cs = compute_lookup.get((src_dev, src_port), {})
        if not cs:
            _ml = _opr.match(r'(swp\d+s)(\d+)', src_port)
            if _ml:
                partner = {'0':'1','1':'0','2':'3','3':'2'}.get(_ml.group(2))
                if partner:
                    cs = compute_lookup.get((src_dev, f"{_ml.group(1)}{partner}"), {})
        if cs:
            t0_lbl  = cs.get('lr', t0_lbl)
            t0_rack = cs.get('t0_rack', t0_rack)
        elif comp_cs:
            t0_lbl  = comp_cs.get('lr', t0_lbl)
            t0_rack = comp_cs.get('t0_rack', t0_rack)
        _ru2 = _opr.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _ru2: t0_rack, t0_elev = _ru2.group(1), _ru2.group(2)
        is_flat    = '-40' in rx_power
        is_missing = 'missing' in rx_power.lower() and '-40' not in rx_power
        if is_missing: continue
        _cm = _opr.match(r'slot\d+/port\d+-(\d+)', remote_port)
        comp_is_phys = (int(_cm.group(1)) % 2 == 1) if _cm else True
        import re as _nrx
        _rux = _nrx.match(r'(Rack \d+) (U\d+)', t0_rack)
        if _rux: t0_rack, t0_elev = _rux.group(1), _rux.group(2)
        rows.append({
            'src_dev': src_dev, 'src_port': src_port, 't0_lbl': t0_lbl, 'is_phys': is_phys,
            'rack': t0_rack, 'elev': t0_elev,
            'remote_dev': remote_dev, 'remote_port': remote_port,
            'rx_power': rx_power, 'is_flat': is_flat, 'is_missing': is_missing,
            'nic_pos':    comp_cs.get('nic_pos', '') or cs.get('nic_pos', ''),
            'ohr':        comp_cs.get('ohr',     '') or cs.get('ohr',     ''),
            'fdf':        comp_cs.get('fdf',     '') or cs.get('fdf',     ''),
            't0_pp':      comp_cs.get('t0_pp',   '') or cs.get('t0_pp',   ''),
            'comp_rack':  comp_rack_num, 'comp_u': comp_u,
            'comp_is_phys': comp_is_phys,
        })
    if not rows: return
    ws = wb_out.create_sheet("Compute Optics")
    ws.sheet_properties.tabColor = "7030A0"
    headers = [
        ("Interface", HDR_BG), ("L&R", HDR_BG), ("T0 Rack", HDR_BG), ("Elevation", HDR_BG),
        ("OHR", "7F6000"), ("FDF", "375623"), ("T0 PP", "17375E"), ("Rx Power", "7030A0"),
        ("Compute Host", "375623"), ("Compute Port", "375623"), ("NIC Position", "833C00"),
        ("Compute Rack", "375623"), ("Compute U", "375623"), ("Flag", "595959"),
    ]
    widths = [12, 6, 12, 6, 28, 28, 30, 30, 40, 16, 14, 12, 8, 26]
    write_header_row(ws, headers, widths)
    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        p = rd['is_phys']; bg = "FFFFFF"; lr_bg = LR_BG if p else LR_LOG
        if rd.get('is_missing'):    flag_txt = "⚠️ Transceiver not responding"
        elif rd.get('is_flat'):      flag_txt = "⬇️ Likely downlink (-40dBm)"
        else:                        flag_txt = ""
        flag_bg = "FFE0B2" if rd.get('is_missing') else ("C8C8C8" if rd.get('is_flat') else "FFFFFF")
        rx_bg   = "FFE0B2" if rd.get('is_missing') else ("C8C8C8" if rd.get('is_flat') else "EAD1F8")
        comp_p  = rd.get('comp_is_phys', True)
        comp_port_bg = "E2F0D9" if comp_p else "D5F5E3"
        vals = [rd['src_port'], rd['t0_lbl'], rd['rack'], rd['elev'],
                rd.get('ohr',''), rd.get('fdf',''), rd.get('t0_pp',''),
                rd['rx_power'], rd['remote_dev'], rd['remote_port'], rd.get('nic_pos',''),
                rd.get('comp_rack',''), rd.get('comp_u',''), flag_txt]
        bgs  = [bg, lr_bg, bg, bg, "FFF2CC","E2F0D9","D9EAF7",
                rx_bg, "E2F0D9", comp_port_bg, "FDDCB5",
                "E2F0D9","E2F0D9", flag_bg]
        for col, (val, cell_bg) in enumerate(zip(vals, bgs), start=1):
            c = ws.cell(r_idx, col); c.value = val; c.fill = fill(cell_bg)
            txt_fg = "888888" if (rd.get('is_flat') or rd.get('is_missing')) else "000000"
            bold = col==2 or (col==len(vals) and bool(flag_txt))
            c.font = Font(bold=bold, color=txt_fg, name="Arial", size=9)
            c.alignment = center()
    print(f"  Compute Optics — {len(rows)} rows ({len(ghost_hosts)} ghost host(s) excluded)")
    draw_compute_borders(ws, host_col=9, port_col=10)

def build_compute_summary(wb_out, compute_real, compute_ghost, ghost_hosts,
                          opt_rows, report_name, fec_rows=None):
    fec_rows = fec_rows or []
    from datetime import datetime
    ws = wb_out.create_sheet("Summary", 0)
    ws.sheet_properties.tabColor = "1F4E79"
    NAVY="1F4E79"; WHITE="FFFFFF"; RED="C00000"; GREEN="1E8449"
    AMBER="B7770D"; GREY="595959"; LGRY="F2F2F2"
    def fill(h):    return PatternFill("solid", fgColor=h)
    def center():   return Alignment(horizontal="center", vertical="center")
    def left():     return Alignment(horizontal="left",   vertical="center")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 16
    for ltr, w in zip("CDEFGHI", [14,14,14,14,14,14,14]):
        ws.column_dimensions[ltr].width = w
    ws.merge_cells("B1:H1"); c=ws["B1"]
    c.value = "COMPUTE LINK VALIDATION — SUMMARY"
    c.fill=fill(NAVY); c.font=Font(bold=True,color=WHITE,name="Arial",size=13)
    c.alignment=center(); ws.row_dimensions[1].height=28
    ws.merge_cells("B2:H2"); c=ws["B2"]
    c.value = f"Report: {report_name}   |   Generated: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.fill=fill("0D7377"); c.font=Font(italic=True,color=WHITE,name="Arial",size=9)
    c.alignment=center(); ws.row_dimensions[2].height=16
    ws.row_dimensions[3].height=8
    flat_count = sum(1 for r in opt_rows if r.get('is_flat'))
    total_issues = len(compute_real) + len(opt_rows) + len(fec_rows)
    kpi = [
        ("TOTAL ISSUES",     total_issues,        NAVY),
        ("HOST LINK ERRORS", len(compute_real),   "375623"),
        ("GHOST LINKS",      len(compute_ghost),   GREY),
        ("COMPUTE OPTICS",   len(opt_rows),         "7030A0"),
        ("FEC ERRORS",       len(fec_rows),         "0070C0"),
        ("⬇️ -40dBm FLAGS",  flat_count,            "808080"),
    ]
    for i, (lbl, val, bg) in enumerate(kpi):
        col = i+2
        ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=30
        c=ws.cell(4,col); c.value=lbl; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=8); c.alignment=center()
        c=ws.cell(5,col); c.value=val; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=20); c.alignment=center()
    ws.row_dimensions[6].height=10
    ws.merge_cells("B7:H7"); c=ws["B7"]
    c.value = f"GHOST COMPUTE TRAYS  ({GHOST_THRESHOLD}+ LLDP errors — not yet online)"
    c.fill=fill(GREY); c.font=Font(bold=True,color=WHITE,name="Arial",size=10)
    c.alignment=center(); ws.row_dimensions[7].height=20
    sub_hdrs = ["Compute Host", "Error Count", "Rack / U", "Status"]
    sub_bgs  = [NAVY, NAVY, NAVY, NAVY]
    for i,(h,bg) in enumerate(zip(sub_hdrs,sub_bgs)):
        c=ws.cell(8,i+2); c.value=h; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=9); c.alignment=center()
    ws.row_dimensions[8].height=16
    ghost_counts = Counter(rd['exp_host'] for rd in compute_ghost)
    ghost_rack_map = {}
    for rd in compute_ghost:
        h = rd.get('exp_host','')
        r = rd.get('comp_rack','')
        if h and r: ghost_rack_map[h] = r
    for row_i, host in enumerate(sorted(ghost_hosts)):
        row = 9 + row_i; ws.row_dimensions[row].height=18
        rack_u = ghost_rack_map.get(host, '')
        for col, (val, bg) in enumerate([
            (host,                     LGRY),
            (ghost_counts.get(host,0), LGRY),
            (rack_u,                   LGRY),
            ("Ghost — not online",     "FFE0E0"),
        ], start=2):
            c=ws.cell(row,col); c.value=val; c.fill=fill(bg)
            c.font=Font(name="Arial",size=10,color="000000")
            c.alignment=center() if col>2 else left()
    next_row = 9 + len(ghost_hosts) + 1
    if not ghost_hosts:
        ws.merge_cells(f"B9:H9"); c=ws["B9"]
        c.value="No ghost hosts detected in this report"
        c.fill=fill("E2F0D9"); c.font=Font(italic=True,name="Arial",size=10,color="1E8449")
        c.alignment=center(); ws.row_dimensions[9].height=18
    ws.row_dimensions[next_row].height=10
    ws.merge_cells(f"B{next_row+1}:H{next_row+1}"); c=ws.cell(next_row+1,2)
    c.value="REAL HOST LINK ERRORS BY COMPUTE HOST"
    c.fill=fill("375623"); c.font=Font(bold=True,color=WHITE,name="Arial",size=10)
    c.alignment=center(); ws.row_dimensions[next_row+1].height=20
    for i,(h,bg) in enumerate(zip(["Compute Host","Error Count","Rack / U"],[NAVY,NAVY,NAVY])):
        c=ws.cell(next_row+2,i+2); c.value=h; c.fill=fill(bg)
        c.font=Font(bold=True,color=WHITE,name="Arial",size=9); c.alignment=center()
    ws.row_dimensions[next_row+2].height=16
    real_counts = Counter(rd['exp_host'] for rd in compute_real)
    real_rack_map = {}
    for rd in compute_real:
        h = rd.get('exp_host','')
        r = rd.get('comp_rack','')
        if h and r: real_rack_map[h] = r
    for ri, (host, cnt) in enumerate(sorted(real_counts.items(), key=lambda x:-x[1])):
        row = next_row+3+ri; ws.row_dimensions[row].height=18
        rack_u = real_rack_map.get(host, '')
        for col, (val,bg) in enumerate([(host,LGRY),(cnt,LGRY),(rack_u,LGRY)], start=2):
            c=ws.cell(row,col); c.value=val; c.fill=fill(bg)
            c.font=Font(name="Arial",size=10); c.alignment=center() if col>2 else left()
    ws.freeze_panes="B2"
    print(f"  Summary tab built — {len(ghost_hosts)} ghost host(s), {len(compute_real)} real errors")

def build_compute_fec_sheet(wb_out, ws_src, ghost_hosts, compute_lookup=None):
    import re as _fre
    compute_lookup = compute_lookup or {}
    if not ws_src: return []
    rows = []
    for row in range(2, ws_src.max_row + 1):
        dev   = str(ws_src.cell(row, 1).value or '').strip()
        iface = str(ws_src.cell(row, 2).value or '').strip()
        lanes = str(ws_src.cell(row, 3).value or '').strip()
        issue = str(ws_src.cell(row, 4).value or '').strip()
        if not dev or not issue: continue
        if not is_compute(dev): continue
        if dev in ghost_hosts: continue
        ber_vals = [float(m) for m in _fre.findall(r'raw-ber=([\d.e+-]+)', issue)]
        if not ber_vals or max(ber_vals) < 1e-07: continue
        max_ber = max(ber_vals)
        if max_ber >= 1e-04:   severity = "🔴 Critical"
        elif max_ber >= 1e-05: severity = "🟠 High"
        elif max_ber >= 1e-06: severity = "🟡 Elevated"
        else:                   severity = "🟢 Marginal"
        cs = cs_lookup(compute_lookup, dev, iface)
        _cr = _fre.match(r'Rack (\d+) U(\d+)', cs.get('comp_rack',''))
        comp_rack = f"Rack {_cr.group(1)}" if _cr else ''
        comp_u    = f"U{_cr.group(2)}"     if _cr else ''
        _t0r_raw = cs.get('t0_rack', '')
        _t0r_m = _fre.match(r'(Rack \d+)(?: (U\d+))?', _t0r_raw)
        t0_rack_s = _t0r_m.group(1)           if _t0r_m else ''
        t0_elev_s = _t0r_m.group(2) or ''  if _t0r_m else ''
        rows.append({
            'dev': dev, 'iface': iface, 'lanes': lanes, 'issue': issue,
            'max_ber': max_ber, 'severity': severity,
            't0_iface': cs.get('t0_iface',''), 'lr': cs.get('lr',''),
            't0_rack':  t0_rack_s, 't0_elev': t0_elev_s,
            'ohr':      cs.get('ohr',''),  'fdf': cs.get('fdf',''),
            't0_pp':    cs.get('t0_pp',''), 'nic_pos': cs.get('nic_pos',''),
            'comp_rack': comp_rack, 'comp_u': comp_u,
        })
    if not rows: return []
    ws = wb_out.create_sheet("FEC Errors")
    ws.sheet_properties.tabColor = "0070C0"
    headers = [
        ("T0 Interface", HDR_BG), ("L&R", HDR_BG), ("T0 Rack", HDR_BG), ("T0 Elev", HDR_BG),
        ("OHR", "7F6000"), ("FDF", "375623"), ("T0 PP", "17375E"),
        ("Compute Host", "17375E"), ("Compute Port", "17375E"), ("NIC Position", "833C00"),
        ("Compute Rack", "17375E"), ("Compute U", "17375E"),
        ("Severity", HDR_BG), ("Max BER", HDR_BG), ("Issue Detail", HDR_BG),
    ]
    widths = [14, 6, 14, 6, 28, 28, 30, 40, 18, 14, 12, 8, 14, 10, 60]
    write_header_row(ws, headers, widths)
    rows = sorted(rows, key=lambda r: (
        r.get('dev',''),
        compute_port_group(r.get('dev',''), r.get('iface','')),
        r.get('iface','')
    ))
    for r_idx, rd in enumerate(rows, start=2):
        ws.row_dimensions[r_idx].height = 15
        sev_bg = {"🔴 Critical":"FFCCCC","🟠 High":"FFE0CC",
                  "🟡 Elevated":"FFF2CC","🟢 Marginal":"E2F0D9"}.get(rd['severity'],"FFFFFF")
        vals = [rd['t0_iface'], rd['lr'], rd['t0_rack'], rd.get('t0_elev',''),
                rd['ohr'], rd['fdf'], rd['t0_pp'],
                rd['dev'], rd['iface'], rd['nic_pos'],
                rd['comp_rack'], rd['comp_u'],
                rd['severity'], f"{rd['max_ber']:.2e}", rd['issue']]
        bgs = ["FFFFFF","D9EAF7","FFFFFF","FFFFFF",
               "FFF2CC","E2F0D9","D9EAF7",
               "E2F0D9","E2F0D9","FDDCB5",
               "E2F0D9","E2F0D9",
               sev_bg,"FFFFFF","FFFFFF"]
        for col,(val,bg) in enumerate(zip(vals,bgs),start=1):
            c = ws.cell(r_idx,col); c.value=val; c.fill=fill(bg)
            c.font = Font(name="Arial",size=9,color="000000")
            c.alignment = Alignment(
                horizontal="left" if col in (4,5,6,14) else "center",
                vertical="center")
    draw_compute_borders(ws, host_col=8, port_col=9)
    print(f"  FEC Errors — {len(rows)} rows ({ws_src.max_row-1-len(rows)} below threshold skipped)")
    return rows

# ── Streamlit App ────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="LV Portal Formatter",
        page_icon="🔧",
        layout="wide",
        initial_sidebar_state="expanded"
    )

    st.title("🔧 LV Portal Formatter - QFAB T0 to Host")
    st.markdown("""
    **Web version - no TKinter or windows com dialog, doesnt require python on local machine** 
    
    - Upload **GPU/Compute cutsheet(s)** (required)
    - Upload **LV Portal Validation Export** (required)
    - (Optional) Upload **T0/T1 Installation cutsheet(s)** for full L&R labels + patch panel data in Mispatches/Downlinks/Optics
    - (Optional) Upload **Previous formatted report** for recurring / history flags
    """)

    with st.sidebar:
        st.header("Upload Files")
        gpu_files = st.file_uploader(
            "GPU / Compute Cutsheet(s)  —  multiple allowed",
            type=["xlsx"],
            accept_multiple_files=True,
            help="Cutsheet(s) containing GPU/Compute to T0 mappings (new or old format supported)"
        )
        report_file = st.file_uploader(
            "LV Portal Validation Export (.xlsx)",
            type=["xlsx"],
            accept_multiple_files=False,
            help="The raw LV Portal report containing LLDP Mismatch / Interface Down / Optics / FEC tabs"
        )
        install_files = st.file_uploader(
            "T0/T1 Installation Cutsheet(s)  —  optional",
            type=["xlsx"],
            accept_multiple_files=True,
            help="Installation cutsheet(s) with T0-T1 patch panel data (for L&R labels and full PP matrix)"
        )
        prev_file = st.file_uploader(
            "Previous Formatted Report  —  optional (for history flags)",
            type=["xlsx"],
            accept_multiple_files=False,
            help="A previously generated _formatted.xlsx to detect recurring issues"
        )

        st.divider()
        st.caption("Processing uses the exact same logic as the original script.")
        run_btn = st.button("🚀 Process Report", type="primary", use_container_width=True,
                            disabled=not (gpu_files and report_file))

    if not run_btn:
        st.info("Upload the required files on the left and click **Process Report** to begin.")
        st.stop()

    # ── Processing ───────────────────────────────────────────────────────────
    temp_dir = tempfile.mkdtemp(prefix="lv_portal_")
    log_capture = io.StringIO()

    try:
        with contextlib.redirect_stdout(log_capture):
            print("=" * 60)
            print("  LV Portal Validation Formatter (Streamlit)")
            print("=" * 60)

            # Write uploaded files to temp
            gpu_paths = []
            for i, f in enumerate(gpu_files):
                p = os.path.join(temp_dir, f"gpu_{i}_{f.name}")
                with open(p, "wb") as out_f:
                    out_f.write(f.getbuffer())
                gpu_paths.append(p)
            print(f"  GPU cutsheets: {[os.path.basename(p) for p in gpu_paths]}")

            report_path = os.path.join(temp_dir, f"report_{report_file.name}")
            with open(report_path, "wb") as out_f:
                out_f.write(report_file.getbuffer())

            prev_report_path = None
            if prev_file:
                prev_report_path = os.path.join(temp_dir, f"prev_{prev_file.name}")
                with open(prev_report_path, "wb") as out_f:
                    out_f.write(prev_file.getbuffer())

            install_paths = []
            if install_files:
                for i, f in enumerate(install_files):
                    p = os.path.join(temp_dir, f"install_{i}_{f.name}")
                    with open(p, "wb") as out_f:
                        out_f.write(f.getbuffer())
                    install_paths.append(p)
                print(f"  Installation cutsheets: {[os.path.basename(p) for p in install_paths]}")

            # Load lookups
            print("\nLoading GPU/Compute cutsheet(s)...")
            compute_lookup = build_compute_lookup(gpu_paths)

            t0, t1, t1_rev, t0_to_pp = {}, {}, {}, {}
            if install_paths:
                print("\nLoading T0/T1 installation cutsheet(s)...")
                t0, t1, t1_rev, t0_to_pp = build_lookup(install_paths)

            # Previous report
            prev_miss, prev_down, prev_opt, prev_rack_map = set(), set(), set(), {}
            if prev_report_path:
                print("\nLoading previous report for history comparison...")
                prev_miss, prev_down, prev_opt, prev_rack_map = get_prev_issues_lv(prev_report_path)

            # Load source workbook
            print(f"\nProcessing report: {os.path.basename(report_path)}")
            wb_src = load_workbook(report_path)

            def find_sheet(wb, *patterns):
                for name in wb.sheetnames:
                    for p in patterns:
                        if p.lower() in name.lower():
                            return wb[name]
                return None

            ws_lldp       = find_sheet(wb_src, 'lldp mismatch', 'lldp', 'mismatch')
            ws_iface_down = find_sheet(wb_src, 'interface down', 'interface_down')
            ws_optics     = find_sheet(wb_src, 'optic errors', 'optic')
            ws_comp_fec   = find_sheet(wb_src, 'fec_ber', 'fec')
            ws_comp_opt   = find_sheet(wb_src, 'optic errors', 'optic')

            if not ws_lldp:
                print("ERROR: Could not find LLDP / Mismatch sheet in the report.")
                st.error("Could not find LLDP / Mismatch sheet in the uploaded report.")
                st.stop()

            # Process LLDP (T0-T1)
            print("\nProcessing LLDP Mismatches & Downlinks (T0-T1 links)...")
            miss_rows, down_rows = process_lldp(ws_lldp, t0, t1, t1_rev, t0_to_pp)

            # Process Compute LLDP
            print("\nProcessing Compute (T0,Host) LLDP errors...")
            compute_real, compute_ghost, ghost_hosts = process_compute_lldp(ws_lldp, compute_lookup)

            # Merge Interface Down if present
            if ws_iface_down:
                print("\nProcessing Interface Down tab (compute links)...")
                id_real, id_ghost, id_ghosts = process_interface_down(ws_iface_down, compute_lookup)
                compute_real.extend(id_real)
                compute_ghost.extend(id_ghost)
                ghost_hosts = ghost_hosts | id_ghosts
                if id_real or id_ghost:
                    print(f"  Merged {len(id_real)} real + {len(id_ghost)} ghost rows from Interface Down")

            print(f"\n  → Mismatches: {len(miss_rows)} | Downlinks: {len(down_rows)}")
            print(f"  → Host Link Errors: {len(compute_real)} | Ghost hosts: {len(compute_ghost)}")

            downlink_set = {(rd['host'], rd['iface']) for rd in down_rows}

            # Build output workbook
            print("\nBuilding formatted output workbook...")
            wb_out = Workbook()
            wb_out.remove(wb_out.active)

            if miss_rows:
                build_mispatches_sheet(wb_out, miss_rows, prev_miss=prev_miss, prev_down=prev_down)
            if down_rows:
                build_downlinks_sheet(wb_out, down_rows, prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)

            fec_rows = build_compute_fec_sheet(wb_out, ws_comp_fec, ghost_hosts, compute_lookup)

            # Summary optics list (simplified)
            opt_rows_for_summary = []
            if ws_comp_opt:
                for _r in range(2, ws_comp_opt.max_row + 1):
                    _remote = str(ws_comp_opt.cell(_r, 1).value or '').strip()
                    _rx = str(ws_comp_opt.cell(_r, 6).value or '').strip()
                    if is_compute(_remote) and _remote not in ghost_hosts:
                        opt_rows_for_summary.append({'is_flat': '-40' in _rx})

            build_compute_summary(wb_out, compute_real, compute_ghost, ghost_hosts,
                                  opt_rows_for_summary, os.path.basename(report_path),
                                  fec_rows=fec_rows)

            if compute_real:
                build_compute_sheet(wb_out, compute_real, "Downlinks", "70AD47",
                                    prev_miss=prev_miss, prev_down=prev_down)
            if compute_ghost:
                build_compute_sheet(wb_out, compute_ghost, "Ghost Links", "808080",
                                    prev_miss=prev_miss, prev_down=prev_down)

            build_compute_optics_sheet(wb_out, ws_comp_opt, ghost_hosts, t0, t1, compute_lookup)

            if ws_optics:
                build_optics_sheet(wb_out, ws_optics, t0, t1, t1_rev, downlink_set, t0_to_pp,
                                   prev_miss=prev_miss, prev_down=prev_down, prev_opt=prev_opt)

            # Save to memory
            output_buffer = io.BytesIO()
            wb_out.save(output_buffer)
            output_buffer.seek(0)
            output_bytes = output_buffer.getvalue()

            print("\n" + "=" * 60)
            print("  Processing complete — ready for download")
            print("=" * 60)

    except Exception as e:
        print(f"\nFATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        st.error(f"Processing failed: {e}")
        with st.expander("Full error traceback"):
            st.code(traceback.format_exc())
        shutil.rmtree(temp_dir, ignore_errors=True)
        st.stop()
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)

    # ── Results UI ───────────────────────────────────────────────────────────
    st.success("✅ Report processed successfully!")

    # Quick metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Mispatches (T0-T1)", len(miss_rows))
    c2.metric("Downlinks (T0-T1)", len(down_rows))
    c3.metric("Host Link Errors", len(compute_real))
    c4.metric("Ghost Hosts", len(ghost_hosts))

    st.divider()

    # Download
    base_name = os.path.splitext(report_file.name)[0]
    st.download_button(
        label="📥 Download Formatted Excel Report",
        data=output_bytes,
        file_name=f"{base_name}_formatted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )

    st.caption("The downloaded file contains the same tabs and formatting as the original desktop version "
               "(Summary, Mispatches, Downlinks, Compute Downlinks, Ghost Links, Compute Optics, FEC Errors, Optics).")

    # Log
    with st.expander("📜 Processing Log (detailed)", expanded=False):
        st.text_area("Log output", log_capture.getvalue(), height=400, label_visibility="collapsed")

if __name__ == "__main__":
    main()
